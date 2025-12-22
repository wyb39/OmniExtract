from bs4 import Tag
import pandas as pd
import subprocess
import tempfile
from typing import Literal, Union
from loguru import logger
from mrkdwn_analysis import MarkdownAnalyzer
import statistics
import os
import openpyxl
import xlrd
from bs4 import BeautifulSoup
import requests
import re
from pylatexenc.latex2text import LatexNodes2Text
from articleUtil import ScienceDirectXmlParser, PubMedCentralXmlParser
import chardet
import json


def detect_encoding(file_path):
    with open(file_path, "rb") as f:
        raw_data = f.read()
        result = chardet.detect(raw_data)
        encoding = result["encoding"]
        return encoding if encoding is not None else "utf-8"


def is_valid(field):
    if pd.isnull(field) or field.strip() == "" or "Unnamed" in field or "NaN" in field:
        return False
    else:
        return True


def write_row(row):
    count = 0
    for field in row:
        if is_valid(field):
            count += 1
    return count > 1


def output_table_to_tsv(table, title, file_path, title_path, encoding):
    header = table["header"]
    rows = table["rows"]
    with open(file_path, "a", encoding=encoding) as f:
        f.write("\t".join(header) + "\n")
        for row in rows:
            if write_row(row):
                f.write("\t".join(row) + "\n")
    # write titles
    with open(title_path, "a", encoding=encoding) as f:
        f.write(f"{title}\n")


def _process_sheet(
    sheet, file_name, index, save_path, encoding, is_xlsx=True, verbose=False
):
    """Process a single Excel worksheet, extract data and save as TSV, MD and META files"""
    # Use os.path.splitext to properly handle filename and extension
    base_name = os.path.splitext(file_name)[0]
    tsv_target_path = os.path.join(save_path, f"{base_name}_{index}.tsv")
    md_target_path = os.path.join(save_path, f"{base_name}_{index}.md")
    table_meta_path = os.path.join(save_path, f"{base_name}_{index}.meta")

    # Check if target files exist to avoid overwriting
    for path in [tsv_target_path, md_target_path, table_meta_path]:
        if os.path.exists(path):
            logger.warning(f"File already exists, will be overwritten: {path}")

    with (
        open(tsv_target_path, "w", encoding=encoding) as tsv,
        open(md_target_path, "w", encoding=encoding) as md,
        open(table_meta_path, "w", encoding=encoding) as meta,
    ):
        header_flag = False

        # Get row data according to Excel format
        rows = (
            sheet.iter_rows(min_row=1, values_only=True)
            if is_xlsx
            else sheet.get_rows()
        )

        for row in rows:
            # Process row data in different formats
            if is_xlsx:
                row_values = list(row)
            else:
                row_values = [item.value for item in row]

            row_tuple = list(filter(None, row_values))
            row_str = []

            for item in row_values:
                if item is None:
                    row_str.append("")
                else:
                    row_str.append(
                        str(item).replace("\n", "").replace("\t", "").strip()
                    )

            if not header_flag:
                if 0 < len(row_tuple) < 3 and len(row_values) > 3:
                    meta.write("	".join(row_tuple))
                    md.write("	".join(row_tuple) + "\n")
                elif len(row_tuple) == 0:
                    # Can choose to keep empty rows or skip
                    pass
                else:
                    if verbose:
                        print("	".join(row_str) + "\n")
                    tsv.write("	".join(row_str) + "\n")
                    md.write("|" + "|".join(row_str) + "|\n")
                    md.write("|" + "|".join(["---"] * len(row_str)) + "|\n")
                    header_flag = True
            else:
                if len(row_tuple) == 0:
                    # Can choose to keep empty rows or skip
                    pass
                else:
                    tsv.write("	".join(row_str) + "\n")
                    md.write("|" + "|".join(row_str) + "|\n")


def parse_excel(file_path, save_path, encoding="utf-8", verbose=False):
    """Parse Excel files, supporting .xlsx and .xls formats, and save data as TSV, MD and META files

    Parameters:

    file_path (str): Excel file path
    save_path (str): Directory to save output files
    encoding (str): File encoding, default is utf-8
    verbose (bool): Whether to print detailed information, default is False
    """
    try:
        # Ensure save directory exists
        os.makedirs(save_path, exist_ok=True)

        file_name = os.path.basename(file_path)

        if file_path.lower().endswith(".xlsx"):
            # Process .xlsx files
            try:
                # Add read_only mode to optimize memory usage
                workbook = openpyxl.load_workbook(
                    file_path, read_only=True, data_only=True
                )
                sheet_names = workbook.sheetnames

                for index, sheet_name in enumerate(sheet_names):
                    sheet = workbook[sheet_name]
                    _process_sheet(
                        sheet,
                        file_name,
                        index,
                        save_path,
                        encoding,
                        is_xlsx=True,
                        verbose=verbose,
                    )

                logger.info(f"Successfully processed XLSX file: {file_path}")
            except Exception as e:
                logger.error(f"Error processing XLSX file: {str(e)}")
                raise
        elif file_path.lower().endswith(".xls"):
            # Process .xls files
            try:
                workbook = xlrd.open_workbook(file_path, encoding_override=encoding)
                sheet_names = workbook.sheet_names()

                for index, sheet_name in enumerate(sheet_names):
                    sheet = workbook.sheet_by_name(sheet_name)
                    _process_sheet(
                        sheet,
                        file_name,
                        index,
                        save_path,
                        encoding,
                        is_xlsx=False,
                        verbose=verbose,
                    )

                logger.info(f"Successfully processed XLS file: {file_path}")
            except Exception as e:
                logger.error(f"Error processing XLS file: {str(e)}")
                raise
        else:
            raise ValueError(
                f"Unsupported file format: {file_path}. Only .xlsx and .xls files are supported."
            )

    except Exception as e:
        logger.error(f"Error parsing Excel file: {str(e)}")
        raise


def parse_csv(file_path, save_path, encoding="utf-8"):
    try:
        # Ensure save directory exists
        os.makedirs(save_path, exist_ok=True)

        file_name = os.path.basename(file_path)
        tsv_target_path = os.path.join(
            save_path, ".".join(file_name.split(".")[:-1]) + "_0.tsv"
        )
        md_target_path = os.path.join(
            save_path, ".".join(file_name.split(".")[:-1]) + "_0.md"
        )
        table_meta_path = os.path.join(
            save_path, ".".join(file_name.split(".")[:-1]) + "_0.meta"
        )

        # Check if target files exist to avoid overwriting
        for path in [tsv_target_path, md_target_path, table_meta_path]:
            if os.path.exists(path):
                logger.warning(f"File already exists, will be overwritten: {path}")

        sep = "\t" if file_path.endswith(".tsv") else ","

        # Read CSV/TSV file
        try:
            encoding = detect_encoding(file_path)
            df = pd.read_csv(file_path, sep=sep, header=None, encoding=encoding)
        except pd.errors.EmptyDataError:
            logger.error(f"Empty file: {file_path}")
            raise
        except pd.errors.ParserError:
            logger.error(
                f"Error parsing file: {file_path}. Check if the file format is correct."
            )
            raise
        except Exception as e:
            logger.error(f"Error reading file: {file_path}. Error: {str(e)}")
            raise

        with (
            open(tsv_target_path, "w", encoding=encoding) as tsv,
            open(md_target_path, "w", encoding=encoding) as md,
            open(table_meta_path, "w", encoding=encoding) as meta,
        ):
            header_flag = False
            for index, row in df.iterrows():
                row_tuple = list(filter(None, list(row)))
                row_str = []
                for item in row:
                    if item is None:
                        row_str.append("")
                    else:
                        row_str.append(str(item).replace("\n", "").replace("\t", ""))
                if not header_flag:
                    if 0 < len(row_tuple) < 3 and len(row) > 3:
                        meta.write("\t".join(row_tuple))
                        md.write("\t".join(row_tuple) + "\n")
                    elif len(row_tuple) == 0:
                        pass
                    else:
                        tsv.write("\t".join(row_str) + "\n")
                        md.write("|" + "|".join(row_str) + "|\n")
                        md.write("|" + "|".join(["---"] * len(row_str)) + "|\n")
                        header_flag = True
                else:
                    if len(row_tuple) == 0:
                        pass
                    else:
                        tsv.write("\t".join(row_str) + "\n")
                        md.write("|" + "|".join(row_str) + "|\n")

        logger.info(f"Successfully processed CSV/TSV file: {file_path}")

    except Exception as e:
        logger.error(f"Error parsing CSV/TSV file: {str(e)}")
        raise


def parse_html_table(
    html_source: Union[str, bytes],
    save_path: str,
    encoding: str = "utf-8",
    verbose: bool = False,
):
    """Parse HTML table from file or URL, and save data as TSV, MD and META files

    Parameters:
    html_source (Union[str, bytes]): HTML file path, URL, or HTML content as bytes
    save_path (str): Directory to save output files
    encoding (str): File encoding, default is utf-8
    verbose (bool): Whether to print detailed information, default is False
    """
    try:
        os.makedirs(save_path, exist_ok=True)

        # Determine source type and read content
        if isinstance(html_source, str):
            if os.path.isfile(html_source):
                with open(html_source, "r", encoding=encoding) as f:
                    html_content = f.read()
                file_name = os.path.basename(html_source)
            elif html_source.startswith(("http://", "https://")):
                response = requests.get(html_source)
                response.raise_for_status()
                html_content = response.text
                file_name = "html_table"
            else:
                html_content = html_source
                file_name = "html_table"
        elif isinstance(html_source, bytes):
            html_content = html_source.decode(encoding)
            file_name = "html_table"
        else:
            raise TypeError(f"Unsupported type for html_source: {type(html_source)}")

        soup = BeautifulSoup(html_content, "html.parser")
        tables = soup.find_all("table")

        if not tables:
            logger.warning("No tables found in HTML source")
            return

        for index, table in enumerate(tables):
            base_name = os.path.splitext(file_name)[0]
            tsv_target_path = os.path.join(save_path, f"{base_name}_{index}.tsv")
            md_target_path = os.path.join(save_path, f"{base_name}_{index}.md")
            table_meta_path = os.path.join(save_path, f"{base_name}_{index}.meta")

            for path in [tsv_target_path, md_target_path, table_meta_path]:
                if os.path.exists(path):
                    logger.warning(f"File already exists, will be overwritten: {path}")

            # Extract table data
            rows = []
            headers = []
            has_header = False
            assert type(table) is Tag
            if table.find("thead"):
                has_header = True
                thead = table.find("thead")
                assert isinstance(thead, Tag)
                header_rows = thead.find_all("tr")
                for row in header_rows:
                    assert isinstance(row, Tag)
                    cells = row.find_all(["th", "td"])
                    headers.append(
                        [
                            cell.get_text(strip=True)
                            .replace("\n", "")
                            .replace("\t", "")
                            for cell in cells
                        ]
                    )

            # Get all rows (including header if no thead)
            table_rows = table.find_all("tr") if hasattr(table, "find_all") else []
            for row in table_rows:
                assert isinstance(row, Tag)
                cells = row.find_all(["th", "td"])
                row_data = [
                    cell.get_text(strip=True).replace("\n", "").replace("\t", "")
                    for cell in cells
                ]
                rows.append(row_data)

            # If no thead, assume first row is header
            if not has_header and rows:
                headers = [rows[0]]
                rows = rows[1:]

            # Flatten headers if multiple header rows
            flat_headers = []
            if headers:
                for header_row in headers:
                    flat_headers.extend(header_row)
                flat_headers = [h for h in flat_headers if h]

            # Extract table caption if exists
            caption = table.find("caption") if hasattr(table, "find") else None
            caption_text = caption.get_text(strip=True) if caption else ""

            with (
                open(tsv_target_path, "w", encoding=encoding) as tsv,
                open(md_target_path, "w", encoding=encoding) as md,
                open(table_meta_path, "w", encoding=encoding) as meta,
            ):
                header_flag = False
                for row in rows:
                    row_tuple = list(filter(None, row))
                    row_str = []
                    for item in row:
                        if not item:
                            row_str.append("")
                        else:
                            row_str.append(item)

                    if not header_flag and flat_headers:
                        tsv.write("\t".join(flat_headers) + "\n")
                        md.write("|" + "|".join(flat_headers) + "|\n")
                        md.write("|" + "|".join(["---"] * len(flat_headers)) + "|\n")
                        header_flag = True

                    if len(row_tuple) == 0:
                        pass
                    else:
                        # Pad or truncate row to match header length
                        if header_flag:
                            if len(row_str) < len(flat_headers):
                                row_str.extend(
                                    [""] * (len(flat_headers) - len(row_str))
                                )
                            elif len(row_str) > len(flat_headers):
                                row_str = row_str[: len(flat_headers)]

                        tsv.write("\t".join(row_str) + "\n")
                        md.write("|" + "|".join(row_str) + "|\n")

            if caption_text:
                meta.write(f"Caption: {caption_text}\n")

            logger.info(
                f"Successfully processed HTML table {index + 1}/{len(tables)} from: {html_source}"
            )

    except Exception as e:
        logger.error(f"Error parsing HTML table: {str(e)}")
        raise


def parse_xml_table_to_tsv(xml_path, save_path, encoding="utf-8", verbose=False):
    """Parse tables from ScienceDirect XML file and save as TSV files

    Parameters:
    xml_path (str): Path to the ScienceDirect XML file
    save_path (str): Directory to save output TSV files
    encoding (str): File encoding, default is utf-8
    verbose (bool): Whether to print detailed information, default is False
    """
    try:
        # Ensure save directory exists
        os.makedirs(save_path, exist_ok=True)

        # Initialize parser and extract tables
        parser = ScienceDirectXmlParser(xml_path)
        parser.parse()
        all_tables = parser.all_tables

        if not all_tables:
            logger.warning(f"No tables found in XML file: {xml_path}")
            return

        # Process each table
        for table_id, table_html in all_tables.items():
            # Create a temporary HTML file to pass to parse_html_table
            temp_html_path = os.path.join(save_path, f"temp_table_{table_id}.html")
            with open(temp_html_path, "w", encoding=encoding) as f:
                f.write(table_html)

            # Use parse_html_table to convert HTML table to TSV
            parse_html_table(
                temp_html_path, save_path, encoding=encoding, verbose=verbose
            )

            # Remove temporary HTML file
            os.remove(temp_html_path)
            logger.info(f"Processed table {table_id} and removed temporary file")

        logger.info(f"Successfully processed all tables from XML file: {xml_path}")

    except Exception as e:
        logger.error(f"Error parsing XML table to TSV: {str(e)}")
        raise


def parse_pubmed_xml_table_to_tsv(xml_path, save_path, encoding="utf-8", verbose=False):
    """Parse tables from PubMed Central XML file and save as TSV files

    Parameters:
    xml_path (str): Path to the PubMed Central XML file
    save_path (str): Directory to save output TSV files
    encoding (str): File encoding, default is utf-8
    verbose (bool): Whether to print detailed information, default is False
    """
    try:
        # Ensure save directory exists
        os.makedirs(save_path, exist_ok=True)

        # Initialize parser and extract tables
        parser = PubMedCentralXmlParser(xml_path)
        parser.parse()
        all_tables = parser.all_tables

        if not all_tables:
            logger.warning(f"No tables found in XML file: {xml_path}")
            return

        # Process each table
        for table_id, table_html in all_tables.items():
            # Create a temporary HTML file to pass to parse_html_table
            temp_html_path = os.path.join(save_path, f"temp_table_{table_id}.html")
            with open(temp_html_path, "w", encoding=encoding) as f:
                f.write(table_html)

            # Use parse_html_table to convert HTML table to TSV
            parse_html_table(
                temp_html_path, save_path, encoding=encoding, verbose=verbose
            )

            # Remove temporary HTML file
            os.remove(temp_html_path)
            logger.info(f"Processed table {table_id} and removed temporary file")

        logger.info(f"Successfully processed all tables from XML file: {xml_path}")

    except Exception as e:
        logger.error(f"Error parsing PubMed XML table to TSV: {str(e)}")
        raise


def parse_tex_tables_to_tsv(
    tex_path: str, save_path: str, encoding: str = "utf-8", verbose: bool = False
):
    """Parse tables from TeX file and save as TSV files using TeXProcessor for better parsing

    This function combines the power of TeXProcessor for accurate LaTeX table parsing
    with the storage capabilities of parse_latex_table to save tables as TSV files.

    Parameters:
    tex_path (str): Path to the TeX file
    save_path (str): Directory to save output TSV files
    encoding (str): File encoding, default is utf-8
    verbose (bool): Whether to print detailed information, default is False
    """
    try:
        # Import TeXProcessor here to avoid circular imports
        from articleUtil import TeXProcessor

        # Ensure save directory exists
        os.makedirs(save_path, exist_ok=True)

        if not os.path.isfile(tex_path):
            raise FileNotFoundError(f"TeX file not found: {tex_path}")

        # Initialize TeXProcessor
        processor = TeXProcessor(tex_path)

        # Process the TeX file to extract tables
        processor.merge_input_files()
        processor.parse_tables()

        if not processor.tables:
            logger.warning(f"No tables found in TeX file: {tex_path}")
            return

        file_name = os.path.basename(tex_path)
        base_name = os.path.splitext(file_name)[0]

        for index, table_data in enumerate(processor.tables):
            table_content = table_data["content"]
            caption = table_data["caption"]

            # Create file paths
            tsv_target_path = os.path.join(save_path, f"{base_name}_{index}.tsv")
            md_target_path = os.path.join(save_path, f"{base_name}_{index}.md")
            table_meta_path = os.path.join(save_path, f"{base_name}_{index}.meta")

            # Check if target files exist
            for path in [tsv_target_path, md_target_path, table_meta_path]:
                if os.path.exists(path):
                    logger.warning(f"File already exists, will be overwritten: {path}")

            # Extract tabular content from table environment
            tabular_pattern = r"\\begin\{tabular\}(.*?)\\end\{tabular\}"
            tabular_match = re.search(tabular_pattern, table_content, re.DOTALL)

            if not tabular_match:
                logger.warning(f"No tabular environment found in table {index + 1}")
                continue

            tabular_content = tabular_match.group(0)

            # Extract column format
            column_format_match = re.search(
                r"\\begin\{tabular\}\{(.*?)\}", tabular_content
            )
            column_format = column_format_match.group(1) if column_format_match else ""

            # Extract data content
            data_content = re.sub(r"\\begin\{tabular\}\{.*?\}", "", tabular_content)
            data_content = re.sub(r"\\end\{tabular\}", "", data_content)

            # Split rows and process cells
            rows = []
            row_data = re.split(r"\\\\", data_content)

            for row in row_data:
                if not row.strip():
                    continue

                # Clean the row
                row = row.strip()
                if row.startswith("\hline"):
                    row = row[6:].strip()
                if row.endswith("\hline"):
                    row = row[:-6].strip()
                if "\hline" in row:
                    row = row.replace("\hline", "").strip()

                if not row.strip():
                    continue

                # Split cells
                cells = re.split(r"(?<!\\)&", row)
                cleaned_cells = []

                for cell in cells:
                    # Remove LaTeX commands and whitespace
                    cleaned_cell = re.sub(r"\\[a-zA-Z]+(?:\{.*?\})?", "", cell)
                    cleaned_cell = re.sub(r"\\[a-zA-Z]+", "", cleaned_cell)
                    cleaned_cell = cleaned_cell.strip()

                    # Handle special LaTeX characters
                    cleaned_cell = cleaned_cell.replace("\\%", "%")
                    cleaned_cell = cleaned_cell.replace("\\$", "$")
                    cleaned_cell = cleaned_cell.replace("\\#", "#")
                    cleaned_cell = cleaned_cell.replace("\\&", "&")
                    cleaned_cell = cleaned_cell.replace("\\_", "_")
                    cleaned_cell = cleaned_cell.replace("\\{", "{")
                    cleaned_cell = cleaned_cell.replace("\\}", "}")

                    # Convert LaTeX to text
                    text_cell = LatexNodes2Text().latex_to_text(cleaned_cell)
                    cleaned_cells.append(text_cell.replace("\n", "").replace("\t", ""))

                if cleaned_cells:  # Only add non-empty rows
                    rows.append(cleaned_cells)

            if not rows:
                logger.warning(f"No valid rows found in table {index + 1}")
                continue

            # Assume first row is header
            headers = [rows[0]] if rows else []
            data_rows = rows[1:] if len(rows) > 1 else []

            # Flatten and clean headers
            flat_headers = []
            if headers:
                for header_row in headers:
                    flat_headers.extend([h for h in header_row if h and h.strip()])

            # Ensure we have valid headers
            if not flat_headers:
                # Generate default headers if none found
                max_cols = max(len(row) for row in rows) if rows else 0
                flat_headers = [f"Column_{i + 1}" for i in range(max_cols)]
                data_rows = rows  # Use all rows as data

            # Process rows and write to files
            with (
                open(tsv_target_path, "w", encoding=encoding) as tsv,
                open(md_target_path, "w", encoding=encoding) as md,
                open(table_meta_path, "w", encoding=encoding) as meta,
            ):
                # Write headers
                tsv.write("\t".join(flat_headers) + "\n")
                md.write("|" + "|".join(flat_headers) + "|\n")
                md.write("|" + "|".join(["---"] * len(flat_headers)) + "|\n")

                # Write data rows
                for row in data_rows:
                    row_str = []
                    for item in row:
                        if not item:
                            row_str.append("")
                        else:
                            row_str.append(item)

                    # Pad or truncate row to match header length
                    if len(row_str) < len(flat_headers):
                        row_str.extend([""] * (len(flat_headers) - len(row_str)))
                    elif len(row_str) > len(flat_headers):
                        row_str = row_str[: len(flat_headers)]

                    tsv.write("\t".join(row_str) + "\n")
                    md.write("|" + "|".join(row_str) + "|\n")

                # Write metadata
                meta.write(f"Caption: {caption}\n")
                meta.write(f"Column format: {column_format}\n")
                meta.write(f"Table index: {index}\n")
                meta.write(f"Source file: {tex_path}\n")

            if verbose:
                logger.info(
                    f"Processed table {index + 1}/{len(processor.tables)}: {caption}"
                )

        logger.info(
            f"Successfully processed {len(processor.tables)} tables from TeX file: {tex_path}"
        )

    except Exception as e:
        logger.error(f"Error parsing TeX tables: {str(e)}")
        raise


def parse_file_for_table_extraction_pdf(
    file_folder_path: str,
    save_folder_path: str,
    disable_multiprocessing: bool = False,
    encoding: str = "utf-8",
    use_llm: bool = False,
    llm_service: Literal[
        "Gemini", "Google Vertex", "Ollama", "Claude", "OpenAI"
    ] = "OpenAI",
    llm_config: dict | None = None,
):
    if not os.path.exists(file_folder_path):
        raise OSError(f"Folder {file_folder_path} does not exist.")

    if not os.path.exists(save_folder_path):
        try:
            os.makedirs(save_folder_path, exist_ok=True)
        except OSError as e:
            raise OSError(f"Failed to create folder {save_folder_path}: {e}")

    err_files = []

    def _run_marker_cli_for_folder(in_folder: str, out_folder: str, cfg: dict):
        args = [
            "marker",
            in_folder,
            "--output_dir",
            out_folder,
            "--workers",
            "1",
            "--output_format",
            cfg.get("output_format", "markdown"),
        ]
        if cfg.get("disable_multiprocessing", False):
            args.append("--disable_multiprocessing")
        if cfg.get("use_llm", False):
            args.append("--use_llm")
            llm_service = cfg.get("llm_service")
            if llm_service:
                args.extend(["--llm_service", llm_service])
        tmp_cfg_path = None
        if cfg:
            tmp_file = tempfile.NamedTemporaryFile(delete=False, suffix=".json")
            tmp_cfg_path = tmp_file.name
            tmp_file.close()
            with open(tmp_cfg_path, "w", encoding="utf-8") as f:
                import json as _json
                _json.dump(cfg, f)
            args.extend(["--config_json", tmp_cfg_path])
        try:
            completed = subprocess.run(args, stdout=None, stderr=None, text=True, check=True)
            if completed.stdout:
                logger.debug(completed.stdout)
            if completed.stderr:
                logger.debug(completed.stderr)
        finally:
            if tmp_cfg_path and os.path.exists(tmp_cfg_path):
                try:
                    os.unlink(tmp_cfg_path)
                except Exception:
                    pass

    cfg = {
        "output_format": "markdown",
        "disable_multiprocessing": disable_multiprocessing,
    }
    if use_llm:
        if llm_config is None:
            raise ValueError("llm_config cannot be None when use_llm is True")
        cfg["use_llm"] = True
        llm_service_map = {
            "Gemini": "marker.services.gemini.GoogleGeminiService",
            "Google Vertex": "marker.services.vertex.VertexService",
            "Ollama": "marker.services.ollama.OllamaService",
            "Claude": "marker.services.claude.ClaudeService",
            "OpenAI": "marker.services.openai.OpenAIService",
        }
        cfg["llm_service"] = llm_service_map[llm_service]
        llm_config_map = {
            "Gemini": ["gemini_api_key"],
            "Google Vertex": ["vertex_project_id"],
            "Ollama": ["ollama_base_url", "ollama_model"],
            "Claude": ["claude_api_key", "claude_model_name"],
            "OpenAI": ["openai_api_key", "openai_model", "openai_base_url"],
        }
        for key in llm_config_map[llm_service]:
            if key in llm_config:
                cfg[key] = llm_config[key]
            else:
                raise ValueError(
                    f"llm_config must contain {key} when use_llm is True and llm_service is {llm_service}"
                )

    pdf_files = [f for f in os.listdir(file_folder_path) if f.lower().endswith(".pdf")]
    if pdf_files:
        import tempfile as _tempfile
        import shutil as _shutil
        _tmp_pdf_dir = _tempfile.mkdtemp(prefix="omniextract_pdf_")
        try:
            for _f in pdf_files:
                _shutil.copy2(
                    os.path.join(file_folder_path, _f),
                    os.path.join(_tmp_pdf_dir, _f),
                )
            _run_marker_cli_for_folder(_tmp_pdf_dir, save_folder_path, cfg)
        finally:
            try:
                _shutil.rmtree(_tmp_pdf_dir)
            except Exception:
                pass

    for file in os.listdir(file_folder_path):
        file_path = os.path.join(file_folder_path, file)
        save_folder_path_file = os.path.join(
            save_folder_path, ".".join(file.split(".")[:-1])
        )
        if not os.path.exists(save_folder_path_file):
            os.makedirs(save_folder_path_file, exist_ok=True)
        if file.endswith(".xlsx") or file.endswith(".xls"):
            try:
                parse_excel(file_path, save_folder_path_file, encoding=encoding)
            except Exception as e:
                err_files.append((file, str(e)))
        elif file.endswith(".csv") or file.endswith(".tsv"):
            try:
                parse_csv(file_path, save_folder_path_file, encoding=encoding)
            except Exception as e:
                err_files.append((file, str(e)))
        else:
            try:
                md_path = os.path.join(
                    save_folder_path_file, ".".join(file.split(".")[:-1]) + ".md"
                )
                if os.path.exists(md_path):
                    with open(md_path, "r", encoding=encoding) as f:
                        text = f.read()
                    text = text.replace("<br>", "")
                    with open(md_path, "w", encoding=encoding) as f:
                        f.write(text)
                # images = rendered.images
                # for image in images:
                #     images[image].save(os.path.join(save_folder_path_file, ".".join(file.split(".")[:-1]), image))
                analyzer = MarkdownAnalyzer(md_path)
                headers = analyzer.identify_emphasis()
                headers = sorted(headers, key=lambda x: x["line"], reverse=True)
                results = []
                extra_tables = []
                # filter headers
                for header in headers:
                    print(header)
                    if "table" in header["text"].lower():
                        header["tables"] = []
                        results.append(header)

                for token in analyzer.tokens:
                    if token.type == "table":
                        flag = False
                        line = token.line
                        for header in results:
                            if "table" in "\t".join(token.meta["header"]).lower():
                                flag = True
                                extra_tables.append(
                                    {
                                        "header": token.meta["header"],
                                        "rows": token.meta["rows"],
                                    }
                                )
                                break
                            elif line > header["line"]:
                                flag = True
                                header["tables"].append(
                                    {
                                        "header": token.meta["header"],
                                        "rows": token.meta["rows"],
                                    }
                                )
                                break
                        if not flag:
                            extra_tables.append(
                                {
                                    "line": line,
                                    "header": token.meta["header"],
                                    "rows": token.meta["rows"],
                                }
                            )

                # filter abnormal tables
                for index, header in enumerate(results):
                    table_lengths = [len(table["header"]) for table in header["tables"]]
                    if len(table_lengths) == 0:
                        continue
                    standard_length = statistics.mode(table_lengths)
                    # Create a new list with filtered tables to avoid modifying during iteration
                    header["tables"] = [
                        table
                        for table in header["tables"]
                        if len(table["header"]) == standard_length
                    ]
                    title = header["text"]
                    tables = header["tables"]
                    for table in tables:
                        tsv_target_path = os.path.join(
                            save_folder_path_file,
                            ".".join(file.split(".")[:-1]) + "_" + str(index) + ".tsv",
                        )
                        table_meta_path = os.path.join(
                            save_folder_path_file,
                            ".".join(file.split(".")[:-1]) + "_" + str(index) + ".meta",
                        )
                        output_table_to_tsv(
                            table,
                            title,
                            tsv_target_path,
                            table_meta_path,
                            encoding=encoding,
                        )

                current_header = ""
                current_header_length = -1
                extra_results = [{"title": current_header, "tables": []}]
                current_index = -1
                for table in extra_tables:
                    new_table = False
                    if len(table["header"]) != current_header_length:
                        current_header = ""
                        current_header_length = len(table["header"])
                        current_index += 1
                        new_table = True
                    min_header_line = table["line"] - 3
                    target_tokens = [
                        item
                        for item in analyzer.tokens
                        if item.line >= min_header_line and item.line <= table["line"]
                    ]
                    for token in target_tokens:
                        if token.content and "table" in token.content.lower():
                            current_header = token.content
                            if new_table:
                                pass
                            else:
                                current_index += 1
                            break
                    if current_index < len(extra_results):
                        extra_results[current_index]["tables"].append(
                            {"header": table["header"], "rows": table["rows"]}
                        )
                        extra_results[current_index]["title"] = current_header
                    else:
                        extra_results.append(
                            {
                                "title": current_header,
                                "tables": [
                                    {"header": table["header"], "rows": table["rows"]}
                                ],
                            }
                        )

                for index, item in enumerate(extra_results):
                    title = item["title"]
                    tables = item["tables"]
                    for table in tables:
                        tsv_target_path = os.path.join(
                            save_folder_path_file,
                            ".".join(file.split(".")[:-1])
                            + "_extra_"
                            + str(index)
                            + ".tsv",
                        )
                        table_meta_path = os.path.join(
                            save_folder_path_file,
                            ".".join(file.split(".")[:-1])
                            + "_extra_"
                            + str(index)
                            + ".meta",
                        )
                        output_table_to_tsv(
                            table,
                            title,
                            tsv_target_path,
                            table_meta_path,
                            encoding=encoding,
                        )

            except Exception as e:
                logger.info(e)
                err_files.append((file, str(e)))
    return err_files


def parse_file_for_table_extraction_tex(
    file_folder_path: str,
    save_folder_path: str,
    encoding: str = "utf-8",
    verbose: bool = False,
):
    """Parse files from TeX files for table extraction, specifically handling TEX files

    Parameters:
    file_folder_path (str): Folder containing TeX files
    save_folder_path (str): Directory to save output TSV files
    encoding (str): File encoding, default is utf-8
    verbose (bool): Whether to print detailed information, default is False
    """
    # Check if input folder exists
    if not os.path.exists(file_folder_path):
        raise OSError(f"Folder {file_folder_path} does not exist.")

    # Ensure output folder exists
    if not os.path.exists(save_folder_path):
        try:
            os.makedirs(save_folder_path, exist_ok=True)
        except OSError as e:
            raise OSError(f"Failed to create folder {save_folder_path}: {e}")

    err_files = []

    # Process each file in the input folder
    for file in os.listdir(file_folder_path):
        file_path = os.path.join(file_folder_path, file)
        save_folder_path_file = os.path.join(
            save_folder_path, ".".join(file.split(".")[:-1])
        )

        # Create subfolder for current file
        if not os.path.exists(save_folder_path_file):
            os.makedirs(save_folder_path_file, exist_ok=True)

        # Handle TEX files specifically
        if file.endswith(".tex"):
            try:
                # Call the TeX table parser
                parse_tex_tables_to_tsv(
                    file_path, save_folder_path_file, encoding=encoding, verbose=verbose
                )
            except Exception as e:
                err_files.append((file, str(e)))
        else:
            # For non-TEX files, add to error list
            err_files.append(
                (
                    file,
                    f"Unsupported file format: {file}. Only TEX files are supported.",
                )
            )

    return err_files


def parse_file_for_table_extraction_pmc(
    file_folder_path: str,
    save_folder_path: str,
    encoding: str = "utf-8",
    verbose: bool = False,
):
    """Parse files from PubMed Central for table extraction, specifically handling XML files

    Parameters:
    file_folder_path (str): Folder containing PubMed Central files
    save_folder_path (str): Directory to save output TSV files
    encoding (str): File encoding, default is utf-8
    verbose (bool): Whether to print detailed information, default is False
    """
    # Check if input folder exists
    if not os.path.exists(file_folder_path):
        raise OSError(f"Folder {file_folder_path} does not exist.")

    # Ensure output folder exists
    if not os.path.exists(save_folder_path):
        try:
            os.makedirs(save_folder_path, exist_ok=True)
        except OSError as e:
            raise OSError(f"Failed to create folder {save_folder_path}: {e}")

    err_files = []

    # Process each file in the input folder
    for file in os.listdir(file_folder_path):
        file_path = os.path.join(file_folder_path, file)
        save_folder_path_file = os.path.join(
            save_folder_path, ".".join(file.split(".")[:-1])
        )

        # Create subfolder for current file
        if not os.path.exists(save_folder_path_file):
            os.makedirs(save_folder_path_file, exist_ok=True)

        # Handle XML files specifically
        if file.endswith(".xlsx") or file.endswith(".xls"):
            try:
                parse_excel(file_path, save_folder_path_file, encoding=encoding)
            except Exception as e:
                err_files.append((file, str(e)))
        elif file.endswith(".csv") or file.endswith(".tsv"):
            try:
                parse_csv(file_path, save_folder_path_file, encoding=encoding)
            except Exception as e:
                err_files.append((file, str(e)))
        elif file.lower().endswith(".xml"):
            try:
                # Call the PubMed XML table parser
                parse_pubmed_xml_table_to_tsv(
                    file_path, save_folder_path_file, encoding=encoding, verbose=verbose
                )
            except Exception as e:
                err_files.append((file, str(e)))
        else:
            # For non-XML files, add to error list
            err_files.append(
                (
                    file,
                    f"Unsupported file format: {file}. Only XML files are supported.",
                )
            )

    return err_files


def parse_file_for_table_extraction_science_direct(
    file_folder_path: str,
    save_folder_path: str,
    encoding: str = "utf-8",
    verbose: bool = False,
):
    """Parse files from ScienceDirect for table extraction, specifically handling XML files

    Parameters:
    file_folder_path (str): Folder containing ScienceDirect files
    save_folder_path (str): Directory to save output TSV files
    encoding (str): File encoding, default is utf-8
    verbose (bool): Whether to print detailed information, default is False
    """
    # Check if input folder exists
    if not os.path.exists(file_folder_path):
        raise OSError(f"Folder {file_folder_path} does not exist.")

    # Ensure output folder exists
    if not os.path.exists(save_folder_path):
        try:
            os.makedirs(save_folder_path, exist_ok=True)
        except OSError as e:
            raise OSError(f"Failed to create folder {save_folder_path}: {e}")

    err_files = []

    # Process each file in the input folder
    for file in os.listdir(file_folder_path):
        file_path = os.path.join(file_folder_path, file)
        save_folder_path_file = os.path.join(
            save_folder_path, ".".join(file.split(".")[:-1])
        )

        # Create subfolder for current file
        if not os.path.exists(save_folder_path_file):
            os.makedirs(save_folder_path_file, exist_ok=True)

        # Handle XML files specifically
        if file.endswith(".xlsx") or file.endswith(".xls"):
            try:
                parse_excel(file_path, save_folder_path_file, encoding=encoding)
            except Exception as e:
                err_files.append((file, str(e)))
        elif file.endswith(".csv") or file.endswith(".tsv"):
            try:
                parse_csv(file_path, save_folder_path_file, encoding=encoding)
            except Exception as e:
                err_files.append((file, str(e)))
        elif file.lower().endswith(".xml"):
            try:
                # Call the XML table parser
                parse_xml_table_to_tsv(
                    file_path, save_folder_path_file, encoding=encoding, verbose=verbose
                )
            except Exception as e:
                err_files.append((file, str(e)))
        else:
            # For non-XML files, add to error list
            err_files.append(
                (
                    file,
                    f"Unsupported file format: {file}. Only XML files are supported.",
                )
            )

    return err_files


def extract_tsv_content_to_json(
    directory_path,
    encoding="utf-8",
    processing_mode: Literal["classify", "example"] = "classify",
    save_path: str = None,
):
    """
    Read all TSV files in a directory, extract content based on processing_mode,
    read the corresponding .meta file (if it exists, otherwise empty string),
    and save the content to a JSON file in a 'json' subdirectory.
    Format: {"table_content": "xxxxx", "table_meta": "xxxx"}

    Args:
        directory_path (str): The directory containing the TSV files.
        encoding (str): Encoding to use for reading files. Default is 'utf-8'.
        processing_mode (str): Mode for processing table content.
                               'classify': Extract first 10 and last 10 rows if > 20 rows.
                               'example': Extract first 10 rows if > 10 rows.
        save_path (str): Directory to save the JSON files. If None, defaults to 'json' subdirectory.
    """
    if not os.path.exists(directory_path):
        logger.error(f"Directory not found: {directory_path}")
        return

    if save_path:
        json_dir = save_path
    else:
        json_dir = os.path.join(directory_path, "json")
    try:
        os.makedirs(json_dir, exist_ok=True)
    except OSError as e:
        logger.error(f"Failed to create directory {json_dir}: {e}")
        return

    for filename in os.listdir(directory_path):
        if filename.lower().endswith(".tsv"):
            tsv_path = os.path.join(directory_path, filename)
            # Assuming meta file has the same base name but .meta extension
            # If tsv is "file.tsv", meta is "file.meta"
            meta_path = os.path.splitext(tsv_path)[0] + ".meta"

            # Read TSV content
            table_content = ""
            try:
                tsv_encoding = detect_encoding(tsv_path)
                file_df = pd.read_csv(
                    tsv_path,
                    sep="\t",
                    encoding=tsv_encoding,
                    header=None,
                )
                
                if processing_mode == "classify":
                    if len(file_df) > 20:
                        for index, row in file_df[0:10].iterrows():
                            row_values = [
                                str(value) if not pd.isnull(value) else " "
                                for value in row.values
                            ]
                            table_content += "\t".join(row_values) + "\n"
                        for index, row in file_df[-10:-1].iterrows():
                            row_values = [
                                str(value) if not pd.isnull(value) else " "
                                for value in row.values
                            ]
                            table_content += "\t".join(row_values) + "\n"
                    else:
                        for index, row in file_df.iterrows():
                            row_values = [
                                str(value) if not pd.isnull(value) else " "
                                for value in row.values
                            ]
                            table_content += "\t".join(row_values) + "\n"
                elif processing_mode == "example":
                    if len(file_df) > 10:
                        for index, row in file_df[0:10].iterrows():
                            row_values = [
                                str(value) if not pd.isnull(value) else " "
                                for value in row.values
                            ]
                            table_content += "\t".join(row_values) + "\n"
                    else:
                        for index, row in file_df.iterrows():
                            row_values = [
                                str(value) if not pd.isnull(value) else " "
                                for value in row.values
                            ]
                            table_content += "\t".join(row_values) + "\n"
                else:
                     logger.warning(f"Unsupported processing mode: {processing_mode}. Using 'classify' mode default.")
                     if len(file_df) > 20:
                        for index, row in file_df[0:10].iterrows():
                            row_values = [
                                str(value) if not pd.isnull(value) else " "
                                for value in row.values
                            ]
                            table_content += "\t".join(row_values) + "\n"
                        for index, row in file_df[-10:-1].iterrows():
                            row_values = [
                                str(value) if not pd.isnull(value) else " "
                                for value in row.values
                            ]
                            table_content += "\t".join(row_values) + "\n"
                     else:
                        for index, row in file_df.iterrows():
                            row_values = [
                                str(value) if not pd.isnull(value) else " "
                                for value in row.values
                            ]
                            table_content += "\t".join(row_values) + "\n"

            except Exception as e:
                logger.error(f"Error reading TSV file {tsv_path}: {e}")
                continue

            # Read Meta content
            table_meta = ""
            if os.path.exists(meta_path):
                try:
                    meta_encoding = detect_encoding(meta_path)
                    with open(meta_path, "r", encoding=meta_encoding) as f:
                        table_meta = f.read()
                except Exception as e:
                    logger.warning(f"Error reading meta file {meta_path}: {e}")
            
            # Create JSON object
            data = {
                "table_content": table_content,
                "table_meta": table_meta
            }
            
            # Save to JSON file
            # Filename for json: "file.tsv" -> "file.json"
            json_filename = os.path.splitext(filename)[0] + ".json"
            json_path = os.path.join(json_dir, json_filename)
            
            try:
                with open(json_path, "w", encoding=encoding) as f:
                    json.dump(data, f, ensure_ascii=False, indent=4)
            except Exception as e:
                logger.error(f"Error writing JSON file {json_path}: {e}")
